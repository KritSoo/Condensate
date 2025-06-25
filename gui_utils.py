"""Utility functions for GUI application."""

import csv
import pandas as pd
import numpy as np
from datetime import datetime
from tkinter import messagebox, filedialog
import tkinter as ttk
import matplotlib.font_manager as fm
import matplotlib as mpl

from gui_config import *
from data_analyzer import (
    calculate_advanced_statistics, format_statistics_for_display,
    detect_anomalies, analyze_trend, add_trend_line_to_plot
)

def configure_thai_font():
    """Configure matplotlib to use a font that supports Thai characters."""
    # List of fonts that typically support Thai characters
    thai_fonts = ['Tahoma', 'Arial Unicode MS', 'TH Sarabun New', 'Browallia New',
                 'Angsana New', 'Microsoft Sans Serif', 'Leelawadee', 'Segoe UI']
    
    # Try to find a font that supports Thai
    for font_name in thai_fonts:
        try:
            # Check if the font exists in the system
            font_path = fm.findfont(fm.FontProperties(family=font_name))
            if font_path and 'ttf' in font_path.lower():
                # Set as the default font
                mpl.rcParams['font.family'] = font_name
                print(f"Using Thai-compatible font: {font_name}")
                return font_name
        except:
            continue
    
    # Fallback to a system default sans-serif font
    mpl.rcParams['font.family'] = 'sans-serif'
    print("Could not find Thai-compatible font, using system default")
    return None

# Global cache for storing data to reduce CSV reading operations
_data_cache = {
    'all_dates': None,  # Cache for available dates
    'data_by_date': {},  # Cache for data by date {date_str: (timestamps, conductivities, temperatures, unit)}
    'last_update_time': None,  # Last time the CSV was modified
    'csv_file_size': 0,  # Size of CSV file to detect changes
}

def clear_cache():
    """Clear the data cache to force fresh data load"""
    global _data_cache
    _data_cache = {
        'all_dates': None,
        'data_by_date': {},
        'last_update_time': None,
        'csv_file_size': 0,
    }

def _is_csv_changed():
    """Check if CSV file has been modified since last read"""
    try:
        import os
        current_size = os.path.getsize('sension7_data.csv')
        current_mtime = os.path.getmtime('sension7_data.csv')
        
        if (_data_cache['csv_file_size'] != current_size or 
            (_data_cache['last_update_time'] is not None and 
             _data_cache['last_update_time'] != current_mtime)):
            _data_cache['csv_file_size'] = current_size
            _data_cache['last_update_time'] = current_mtime
            return True
        return False
    except Exception as e:
        print(f"Error checking CSV status: {e}")
        return True  # Assume it changed if we can't check

def get_available_dates(force_refresh=False):
    """Get list of available dates from CSV file with caching."""
    global _data_cache
    
    # Return cached data if available and CSV hasn't changed
    if not force_refresh and _data_cache['all_dates'] is not None and not _is_csv_changed():
        return _data_cache['all_dates']
        
    available_dates = set()
    try:
        with open('sension7_data.csv', 'r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                timestamp = datetime.strptime(row['Timestamp'], '%Y-%m-%d %H:%M:%S')
                available_dates.add(timestamp.strftime('%Y-%m-%d'))
        
        dates_list = sorted(list(available_dates))
        _data_cache['all_dates'] = dates_list
        return dates_list
    except Exception as e:
        print(f"Error reading dates from CSV: {e}")
        return []

def calculate_statistics(data):
    """Calculate statistics from conductivity data."""
    try:
        # Convert to list if it's a pandas Series or DataFrame
        if hasattr(data, 'values'):
            data_list = data.values.tolist()
        else:
            data_list = list(data)
            
        # Use advanced statistics function
        return calculate_advanced_statistics(data_list)
    except Exception as e:
        print(f"Error in calculate_statistics: {e}")
        # Return basic statistics as fallback
        return {
            'min': data.min() if hasattr(data, 'min') else min(data),
            'max': data.max() if hasattr(data, 'max') else max(data),
            'mean': data.mean() if hasattr(data, 'mean') else sum(data) / len(data),
            'std': data.std() if hasattr(data, 'std') else np.std(data),
            'count': len(data)
        }

def update_statistics(values, format_for_display=True):
    """Update statistics display with advanced metrics."""
    try:
        if values and len(values) > 0:
            # Calculate advanced statistics
            stats = calculate_advanced_statistics(values)
            
            if format_for_display:
                # Format for display in GUI
                return format_statistics_for_display(stats)
            else:
                return stats
        
        if format_for_display:
            return {"ค่าสูงสุด": "0.00", "ค่าต่ำสุด": "0.00", "ค่าเฉลี่ย": "0.00"}
        else:
            return {"max": 0.0, "min": 0.0, "mean": 0.0}
    except Exception as e:
        print(f"Error calculating statistics: {e}")
        if format_for_display:
            return {"ค่าสูงสุด": "0.00", "ค่าต่ำสุด": "0.00", "ค่าเฉลี่ย": "0.00"}
        else:
            return {"max": 0.0, "min": 0.0, "mean": 0.0}

def export_to_excel():
    """Export data to Excel file."""
    try:
        from openpyxl import Workbook
        from tkinter import filedialog
        
        # Get save file location
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*")]
        )
        
        if not filename:
            return
            
        # Create workbook and select active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Sensor Data"
        
        # Write headers
        ws['A1'] = 'Timestamp'
        ws['B1'] = 'Conductivity'
        ws['C1'] = 'Unit'
        ws['D1'] = 'Temperature'
        
        # Read and write data
        with open('sension7_data.csv', 'r') as file:
            reader = csv.DictReader(file)
            for idx, row in enumerate(reader, start=2):
                ws[f'A{idx}'] = row['Timestamp']
                ws[f'B{idx}'] = float(row['Conductivity'])
                ws[f'C{idx}'] = row['Unit']
                ws[f'D{idx}'] = float(row['Temperature']) if row['Temperature'] else None
        
        # Save workbook
        wb.save(filename)
        messagebox.showinfo("Success", "Data exported successfully!")
        
    except Exception as e:
        messagebox.showerror("Error", f"Failed to export data: {str(e)}")

def refresh_date_list():
    """Update the date combobox with available dates."""
    global date_combobox
    if date_combobox is None:
        print("Warning: Date combobox not initialized")
        return
    
    try:
        dates = get_available_dates(force_refresh=True)
        date_combobox['values'] = dates
        if dates:
            date_combobox.current(0)
    except Exception as e:
        print(f"Error refreshing date list: {e}")

def read_csv_data(date_str, force_refresh=False):
    """Read data from CSV file for given date with caching."""
    global _data_cache
    
    # Return cached data if available and CSV hasn't changed
    if not force_refresh and date_str in _data_cache['data_by_date'] and not _is_csv_changed():
        return _data_cache['data_by_date'][date_str]
    
    timestamps = []
    conductivities = []
    temperatures = []
    unit = "uS/cm"
    
    try:
        with open('sension7_data.csv', 'r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                timestamp = datetime.strptime(row['Timestamp'], '%Y-%m-%d %H:%M:%S')
                if timestamp.strftime('%Y-%m-%d') == date_str:
                    timestamps.append(timestamp)
                    conductivities.append(float(row['Conductivity']))
                    temperatures.append(float(row['Temperature']) if row['Temperature'] else None)
                    unit = row['Unit']
        
        # Cache the results
        result = (timestamps, conductivities, temperatures, unit)
        _data_cache['data_by_date'][date_str] = result
        return result
        
    except Exception as e:
        print(f"Error reading CSV: {e}")
        return [], [], [], "uS/cm"

def detect_data_anomalies(timestamps, values, method='zscore'):
    """
    ตรวจจับความผิดปกติในข้อมูล
    
    Parameters:
    -----------
    timestamps : list
        รายการวันเวลาที่เก็บข้อมูล
    values : list
        รายการค่าที่วัดได้
    method : str
        วิธีการตรวจจับความผิดปกติ ('zscore', 'iqr', หรือ 'isolation_forest')
        
    Returns:
    --------
    tuple
        (anomaly_indices, anomaly_timestamps, anomaly_values)
    """
    try:
        # ใช้ฟังก์ชันจาก data_analyzer
        return detect_anomalies(timestamps, values, method)
    except Exception as e:
        print(f"Error detecting anomalies: {e}")
        return [], [], []

def get_data_trend_analysis(timestamps, values, window_size=5):
    """
    วิเคราะห์แนวโน้มของข้อมูล
    
    Parameters:
    -----------
    timestamps : list
        รายการวันเวลาที่เก็บข้อมูล
    values : list
        รายการค่าที่วัดได้
    window_size : int
        ขนาดของหน้าต่างสำหรับการคำนวณค่าเฉลี่ยเคลื่อนที่
        
    Returns:
    --------
    dict
        ข้อมูลแนวโน้มของข้อมูล
    """
    try:
        # ใช้ฟังก์ชันจาก data_analyzer
        return analyze_trend(timestamps, values, window_size)
    except Exception as e:
        print(f"Error analyzing trend: {e}")
        return {
            'trend_direction': 'error',
            'trend_strength': 0.0,
            'moving_average': [],
            'slope': 0.0,
            'r_squared': 0.0,
            'p_value': 1.0
        }

def read_csv_data_with_analysis(date_str, force_refresh=False):
    """Read data from CSV file for given date with additional analysis."""
    # อ่านข้อมูลด้วยฟังก์ชันเดิม
    timestamps, conductivities, temperatures, unit = read_csv_data(date_str, force_refresh)
    
    if not timestamps:
        return timestamps, conductivities, temperatures, unit, None, None
    
    # วิเคราะห์ข้อมูลเพิ่มเติม
    try:
        # ตรวจจับความผิดปกติ
        cond_anomalies = detect_data_anomalies(timestamps, conductivities)
        temp_anomalies = detect_data_anomalies(timestamps, temperatures)
        
        # วิเคราะห์แนวโน้ม
        cond_trend = get_data_trend_analysis(timestamps, conductivities)
        temp_trend = get_data_trend_analysis(timestamps, temperatures)
        
        # รวมผลการวิเคราะห์
        analysis = {
            'conductivity': {
                'anomalies': cond_anomalies,
                'trend': cond_trend
            },
            'temperature': {
                'anomalies': temp_anomalies,
                'trend': temp_trend
            }
        }
        
        # คำนวณสถิติขั้นสูง
        cond_stats = calculate_advanced_statistics(conductivities)
        temp_stats = calculate_advanced_statistics(temperatures)
        
        # รวมสถิติ
        statistics = {
            'conductivity': cond_stats,
            'temperature': temp_stats
        }
        
        return timestamps, conductivities, temperatures, unit, analysis, statistics
        
    except Exception as e:
        print(f"Error in data analysis: {e}")
        return timestamps, conductivities, temperatures, unit, None, None
